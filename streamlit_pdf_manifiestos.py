from __future__ import annotations

import io
from datetime import datetime

import streamlit as st
from openpyxl import load_workbook

GENERATOR_IMPORT_ERROR: ModuleNotFoundError | None = None

try:
    from generate_shipping_labels import (
        generate_generic_zip_bytes,
        generate_zip_bytes,
        read_rows,
    )
except ModuleNotFoundError as exc:
    GENERATOR_IMPORT_ERROR = exc
    generate_generic_zip_bytes = None
    generate_zip_bytes = None
    read_rows = None


st.set_page_config(
    page_title="Generador de PDFs por Manifiesto",
    page_icon="📦",
    layout="centered",
)


def get_sheet_names(file_bytes: bytes) -> list[str]:
    workbook = load_workbook(io.BytesIO(file_bytes), read_only=True, data_only=True)
    return workbook.sheetnames


def default_sheet_index(sheet_names: list[str], prefer: str | None = None) -> int:
    """Indice de la hoja por defecto: la que contenga ``prefer`` (ej. 'dian'),
    o la primera si no hay coincidencia."""
    if prefer:
        for index, name in enumerate(sheet_names):
            if prefer in name.strip().lower():
                return index
    return 0


st.title("Generador de etiquetas 4x6")

if GENERATOR_IMPORT_ERROR is not None:
    st.error("Falta una dependencia necesaria para generar los PDFs.")
    st.code("pip install -r requirements.txt")
    st.write("Dependencia reportada:", f"`{GENERATOR_IMPORT_ERROR}`")
    st.stop()

modo = st.radio(
    "Tipo de etiqueta",
    options=["Con marca (Encargomio)", "Generica (simulada)"],
    index=0,
    horizontal=True,
)
generic_mode = modo.startswith("Gen")

if generic_mode:
    st.write(
        "Sube el Excel del manifiesto DIAN y descarga un ZIP con etiquetas "
        "simuladas (sin marca), una por fila. Cada guia mantiene siempre el "
        "mismo diseno."
    )
else:
    st.write("Sube un Excel y descarga un ZIP con un PDF por cada fila del manifiesto.")

uploaded_file = st.file_uploader(
    "Cargar archivo Excel",
    type=["xlsx"],
    accept_multiple_files=False,
)

if uploaded_file is not None:
    file_bytes = uploaded_file.getvalue()
    sheet_names = get_sheet_names(file_bytes)

    sheet_index = default_sheet_index(sheet_names, prefer="dian" if generic_mode else None)
    selected_sheet = st.selectbox(
        "Hoja a procesar",
        options=sheet_names,
        index=sheet_index,
    )

    button_label = "Preparar ZIP generico" if generic_mode else "Preparar ZIP"
    if st.button(button_label, type="primary"):
        with st.spinner("Generando PDFs..."):
            rows = read_rows(io.BytesIO(file_bytes), selected_sheet)

            if not rows:
                st.error("No se encontraron filas validas con numero de Envio o Guia.")
            else:
                timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")

                if generic_mode:
                    zip_bytes, failures = generate_generic_zip_bytes(rows)
                    zip_name = f"etiquetas_genericas_{selected_sheet}_{timestamp}.zip"
                    ok_count = len(rows) - len(failures)
                    st.success(
                        f"Se prepararon {ok_count} etiquetas genericas de {len(rows)} filas."
                    )
                    if failures:
                        st.warning(
                            f"{len(failures)} fila(s) fallaron y se omitieron del ZIP:"
                        )
                        st.table(
                            [
                                {"Guia": f["shipment_number"], "Error": f["error"]}
                                for f in failures
                            ]
                        )
                else:
                    zip_bytes = generate_zip_bytes(rows)
                    zip_name = f"etiquetas_{selected_sheet}_{timestamp}.zip"
                    st.success(f"Se prepararon {len(rows)} etiquetas.")

                st.download_button(
                    label="Descargar ZIP con PDFs",
                    data=zip_bytes,
                    file_name=zip_name,
                    mime="application/zip",
                )
else:
    st.info("Selecciona un archivo Excel para comenzar.")
