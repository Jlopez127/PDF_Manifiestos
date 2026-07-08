from __future__ import annotations

import argparse
import hashlib
import io
import random
import re
import unicodedata
import zipfile
from dataclasses import dataclass
from pathlib import Path
from typing import BinaryIO, Iterable

from openpyxl import load_workbook

REPORTLAB_IMPORT_ERROR: ModuleNotFoundError | None = None

try:
    from reportlab.graphics import renderPDF
    from reportlab.graphics.barcode import code128
    from reportlab.graphics.barcode import createBarcodeDrawing
    from reportlab.lib.colors import HexColor
    from reportlab.pdfbase.pdfmetrics import stringWidth
    from reportlab.pdfgen import canvas
except ModuleNotFoundError as exc:
    REPORTLAB_IMPORT_ERROR = exc
    renderPDF = None
    code128 = None
    createBarcodeDrawing = None
    HexColor = None
    canvas = None

    def stringWidth(*args: object, **kwargs: object) -> float:
        raise ModuleNotFoundError(
            "Falta la dependencia 'reportlab'. Instalala con: pip install reportlab"
        ) from REPORTLAB_IMPORT_ERROR


POINTS_PER_INCH = 72.0
INCH = POINTS_PER_INCH
PAGE_WIDTH = 6 * POINTS_PER_INCH
PAGE_HEIGHT = 4 * POINTS_PER_INCH
PAGE_SIZE = (PAGE_WIDTH, PAGE_HEIGHT)
MARGIN = 0.14 * POINTS_PER_INCH
CONTENT_WIDTH = PAGE_WIDTH - (2 * MARGIN)


def ensure_reportlab() -> None:
    if REPORTLAB_IMPORT_ERROR is not None:
        raise ModuleNotFoundError(
            "Falta la dependencia 'reportlab'. Instalala con: pip install reportlab"
        ) from REPORTLAB_IMPORT_ERROR


def normalize_key(value: str) -> str:
    text = unicodedata.normalize("NFKD", str(value or ""))
    text = "".join(ch for ch in text if not unicodedata.combining(ch))
    text = text.lower().strip()
    text = re.sub(r"[^a-z0-9]+", " ", text)
    return " ".join(text.split())


def pick_first(record: dict[str, object], aliases: Iterable[str], default: str = "") -> str:
    for alias in aliases:
        key = normalize_key(alias)
        if key in record:
            value = record[key]
            if value is None:
                continue
            if isinstance(value, float) and value.is_integer():
                return str(int(value))
            return str(value).strip()
    return default


def to_float(value: str) -> float | None:
    if not value:
        return None
    try:
        return float(str(value).replace(",", "."))
    except ValueError:
        return None


def wrap_text(text: str, font_name: str, font_size: float, max_width: float) -> list[str]:
    words = (text or "").split()
    if not words:
        return []

    lines: list[str] = []
    current = words[0]

    for word in words[1:]:
        test_line = f"{current} {word}"
        if stringWidth(test_line, font_name, font_size) <= max_width:
            current = test_line
            continue
        lines.append(current)
        current = word

    lines.append(current)
    return lines


def draw_wrapped(
    pdf: canvas.Canvas,
    text: str,
    x: float,
    y: float,
    width: float,
    font_name: str = "Helvetica",
    font_size: float = 8.5,
    leading: float = 10.0,
    max_lines: int | None = None,
) -> float:
    lines = wrap_text(text, font_name, font_size, width)
    if max_lines is not None and len(lines) > max_lines:
        lines = lines[:max_lines]
        if lines:
            trimmed = lines[-1]
            while trimmed and stringWidth(trimmed + "...", font_name, font_size) > width:
                trimmed = trimmed[:-1]
            lines[-1] = (trimmed.rstrip() + "...") if trimmed else "..."

    pdf.setFont(font_name, font_size)
    cursor = y
    for line in lines:
        pdf.drawString(x, cursor, line)
        cursor -= leading
    return cursor


@dataclass
class LabelRow:
    shipment_number: str
    shipment_date: str
    sender_name: str
    sender_address: str
    sender_phone: str
    sender_city: str
    sender_state: str
    recipient_name: str
    recipient_address: str
    recipient_phone: str
    recipient_city: str
    recipient_state: str
    content: str
    pieces: str
    weight_lb: str
    weight_kg: str
    declared_value: str
    tariff_code: str
    manifest: str
    instructions: str
    cost: str

    @classmethod
    def from_record(cls, record: dict[str, object]) -> "LabelRow":
        return cls(
            shipment_number=pick_first(record, ["envio", "guia", "guia / consignment", "guia consignment", "numero de guia", "numero guia"]),
            shipment_date=pick_first(record, ["fecha guia", "fecha", "fecha envio"]),
            sender_name=pick_first(record, ["compania remitente", "remitente", "remitente nombre"]),
            sender_address=pick_first(record, ["remitente direccion", "direccion remitente"]),
            sender_phone=pick_first(record, ["remitente telefono", "telefono remitente"]),
            sender_city=pick_first(record, ["remitente ciudad", "ciudad remitente"]),
            sender_state=pick_first(record, ["remitente estado", "estado remitente"]),
            recipient_name=pick_first(record, ["nombre destino", "destinatario", "nombre destinatario"]),
            recipient_address=pick_first(record, ["destino direccion", "direccion destino", "direccion destinatario"]),
            recipient_phone=pick_first(record, ["destino telefono", "telefono destino", "telefono destinatario"]),
            recipient_city=pick_first(record, ["destino ciudad", "ciudad destino"]),
            recipient_state=pick_first(record, ["destino estado", "estado destino"]),
            content=pick_first(record, ["contenido", "descripcion", "producto"]),
            pieces=pick_first(record, ["piezas", "cantidad piezas"], "1"),
            weight_lb=pick_first(record, ["peso libras", "peso lb"]),
            weight_kg=pick_first(record, ["peso kilos", "peso kg"]),
            declared_value=pick_first(record, ["valor declarado", "valor"]),
            tariff_code=pick_first(record, ["posicion arancelaria", "partida arancelaria", "hs code"]),
            manifest=pick_first(record, ["manifiesto"]),
            instructions=pick_first(record, ["instrucciones", "observaciones"]),
            cost=pick_first(record, ["costo", "coste"]),
        )


def read_rows(workbook_source: str | Path | BinaryIO, sheet_name: str | None) -> list[LabelRow]:
    workbook = load_workbook(workbook_source, data_only=True)
    worksheet = workbook[sheet_name] if sheet_name else workbook[workbook.sheetnames[0]]

    headers = [normalize_key(cell.value) for cell in worksheet[1]]
    rows: list[LabelRow] = []

    for row in worksheet.iter_rows(min_row=2, values_only=True):
        if all(value in (None, "") for value in row):
            continue
        record = {
            headers[index]: value
            for index, value in enumerate(row)
            if index < len(headers) and headers[index]
        }
        label = LabelRow.from_record(record)
        if not label.shipment_number:
            continue
        rows.append(label)

    return rows


def format_weight(lb: str, kg: str) -> str:
    lb_value = to_float(lb)
    kg_value = to_float(kg)
    parts: list[str] = []
    if lb_value is not None:
        parts.append(f"{lb_value:.2f} Lb")
    elif lb:
        parts.append(f"{lb} Lb")
    if kg_value is not None:
        parts.append(f"{kg_value:.2f} Kg")
    elif kg:
        parts.append(f"{kg} Kg")
    return " - ".join(parts)


def sanitize_filename(value: str) -> str:
    cleaned = re.sub(r'[<>:"/\\|?*]+', "_", value.strip())
    cleaned = re.sub(r"\s+", "_", cleaned)
    return cleaned[:120] or "etiqueta"


def draw_centered_block(
    pdf: canvas.Canvas,
    center_x: float,
    top_y: float,
    lines: list[tuple[str, str, float, str]],
    leading: float,
) -> float:
    y = top_y
    for text, font_name, font_size, color in lines:
        pdf.setFillColor(HexColor(color))
        pdf.setFont(font_name, font_size)
        pdf.drawCentredString(center_x, y, text)
        y -= leading
    return y


def draw_box_icon(pdf: canvas.Canvas, center_x: float, center_y: float, size: float) -> None:
    half = size / 2
    left = center_x - half
    bottom = center_y - half
    pdf.setStrokeColor(HexColor("#6B7280"))
    pdf.setLineWidth(1)
    pdf.rect(left, bottom, size * 0.72, size * 0.62, stroke=1, fill=0)
    pdf.line(left, bottom + size * 0.62, left + size * 0.22, bottom + size * 0.78)
    pdf.line(left + size * 0.22, bottom + size * 0.78, left + size * 0.72, bottom + size * 0.62)
    pdf.line(left + size * 0.36, bottom + size * 0.70, left + size * 0.36, bottom + size * 0.08)
    pdf.line(left + size * 0.22, bottom + size * 0.78, left + size * 0.48, bottom + size * 0.60)
    pdf.line(left + size * 0.48, bottom + size * 0.60, left + size * 0.48, bottom + size * 0.42)
    pdf.line(left + size * 0.60, bottom + size * 0.68, left + size * 0.86, bottom + size * 0.94)
    pdf.line(left + size * 0.86, bottom + size * 0.94, left + size * 0.72, bottom + size * 0.62)
    pdf.line(left + size * 0.06, bottom + size * 0.16, left + size * 0.16, bottom + size * 0.12)
    pdf.line(left + size * 0.10, bottom + size * 0.22, left + size * 0.20, bottom + size * 0.18)


def draw_warning_icon(pdf: canvas.Canvas, center_x: float, center_y: float, size: float) -> None:
    pdf.setStrokeColor(HexColor("#111827"))
    pdf.setLineWidth(1.2)
    path = pdf.beginPath()
    path.moveTo(center_x, center_y + size * 0.55)
    path.lineTo(center_x - size * 0.52, center_y - size * 0.35)
    path.lineTo(center_x + size * 0.52, center_y - size * 0.35)
    path.close()
    pdf.drawPath(path, stroke=1, fill=0)
    pdf.setFont("Helvetica-Bold", size * 0.65)
    pdf.drawCentredString(center_x, center_y - size * 0.12, "!")


def draw_house_icon(pdf: canvas.Canvas, center_x: float, center_y: float, size: float) -> None:
    left = center_x - size * 0.36
    bottom = center_y - size * 0.34
    pdf.setStrokeColor(HexColor("#6B7280"))
    pdf.setLineWidth(1)
    path = pdf.beginPath()
    path.moveTo(center_x, center_y + size * 0.48)
    path.lineTo(center_x - size * 0.46, center_y)
    path.lineTo(center_x - size * 0.34, center_y)
    path.lineTo(center_x - size * 0.34, bottom)
    path.lineTo(center_x - size * 0.10, bottom)
    path.lineTo(center_x - size * 0.10, bottom + size * 0.24)
    path.lineTo(center_x + size * 0.10, bottom + size * 0.24)
    path.lineTo(center_x + size * 0.10, bottom)
    path.lineTo(center_x + size * 0.34, bottom)
    path.lineTo(center_x + size * 0.34, center_y)
    path.lineTo(center_x + size * 0.46, center_y)
    path.close()
    pdf.drawPath(path, stroke=1, fill=0)


def build_sender_lines(row: LabelRow) -> list[str]:
    lines = [row.sender_name or "N/D"]
    if row.sender_address:
        lines.append(row.sender_address)
    city_line = " ".join(part for part in [row.sender_city, row.sender_state] if part)
    if city_line:
        lines.append(f"{city_line} U.S.A.")
    return lines[:3]


def build_recipient_lines(row: LabelRow) -> list[str]:
    lines = [row.recipient_name or "N/D"]
    if row.recipient_address:
        lines.extend(wrap_text(row.recipient_address, "Helvetica", 7.4, 148)[:2])
    city_line = ", ".join(part for part in [row.recipient_city, row.recipient_state] if part)
    if city_line:
        lines.append(city_line)
    if row.recipient_phone:
        lines.append(f"Tel: {row.recipient_phone}")
    return lines[:4]


def draw_code128_barcode(
    pdf: canvas.Canvas,
    value: str,
    x: float,
    y: float,
    max_width: float,
    height: float,
) -> None:
    ensure_reportlab()
    # Barras lo mas gruesas posible para que la pistola lea mejor: se parte de un
    # grosor objetivo y solo se adelgaza si el codigo no cabe en el ancho
    # disponible (asi nunca se sale del panel).
    bar_width = 2.0
    drawing = createBarcodeDrawing(
        "Code128", value=value, barHeight=height, barWidth=bar_width, humanReadable=False
    )
    guard = 0
    while drawing.width > max_width and drawing.width > 0 and guard < 8:
        bar_width = bar_width * (max_width / drawing.width) * 0.98
        drawing = createBarcodeDrawing(
            "Code128", value=value, barHeight=height, barWidth=bar_width, humanReadable=False
        )
        guard += 1

    # centrado dentro del ancho disponible
    drawing_x = x + (max_width - drawing.width) / 2
    renderPDF.draw(drawing, pdf, drawing_x, y)


def draw_label(pdf: canvas.Canvas, row: LabelRow) -> None:
    ensure_reportlab()
    pdf.setFillColor(HexColor("#FFFFFF"))
    pdf.rect(0, 0, PAGE_WIDTH, PAGE_HEIGHT, stroke=0, fill=1)

    outer_border = HexColor("#6B7280")
    divider = HexColor("#BFC5CD")
    split_x = PAGE_WIDTH * 0.505
    left_x0 = MARGIN
    left_x1 = split_x - 6
    right_x0 = split_x + 8
    right_x1 = PAGE_WIDTH - MARGIN
    left_width = left_x1 - left_x0
    right_width = right_x1 - right_x0

    pdf.setStrokeColor(outer_border)
    pdf.setLineWidth(1)
    pdf.rect(0.8, 0.8, PAGE_WIDTH - 1.6, PAGE_HEIGHT - 1.6, stroke=1, fill=0)
    pdf.line(split_x, 0, split_x, PAGE_HEIGHT)

    barcode_value = row.shipment_number.strip()
    draw_code128_barcode(
        pdf,
        barcode_value,
        left_x0 + 4,
        PAGE_HEIGHT - 66,
        left_width - 8,
        56,
    )
    pdf.setFillColor(HexColor("#111827"))
    pdf.setFont("Helvetica", 8.3)
    pdf.drawCentredString(left_x0 + left_width / 2, PAGE_HEIGHT - 76, barcode_value)

    y = PAGE_HEIGHT - 96
    y = draw_centered_block(
        pdf,
        left_x0 + left_width / 2,
        y,
        [
            ("!Hola!", "Helvetica-Bold", 20, "#F97316"),
            ("Somos tu mejor aliado en", "Helvetica", 9.5, "#111827"),
            ("Compras y Envios en el exterior", "Helvetica", 9.5, "#111827"),
            ("Ideal para Negocios", "Helvetica-Bold", 9.5, "#111827"),
        ],
        leading=11,
    )

    box_top = 116
    box_bottom = 68
    label_col_x = left_x0 + 8
    value_x = left_x0 + 82
    envio_x = left_x1 - 64

    pdf.setStrokeColor(divider)
    pdf.setLineWidth(0.8)

    sender_lines = build_sender_lines(row)
    recipient_lines = build_recipient_lines(row)

    pdf.setFillColor(HexColor("#111827"))
    pdf.setFont("Helvetica-Bold", 10)
    sender_label_y = box_top + 18
    recipient_label_y = box_bottom - 7
    pdf.drawString(label_col_x, sender_label_y, "Remitente:")
    pdf.drawString(label_col_x, recipient_label_y, "Destinatario:")

    pdf.setFont("Helvetica", 6.7)
    sender_text_y = box_top + 21
    sender_text_width = envio_x - value_x - 6
    current_sender_y = sender_text_y
    for line in sender_lines:
        current_sender_y = draw_wrapped(
            pdf,
            line,
            value_x,
            current_sender_y,
            sender_text_width,
            font_name="Helvetica",
            font_size=6.7,
            leading=7.2,
            max_lines=2,
        )
        current_sender_y -= 1.5

    recipient_y = box_bottom - 6
    for idx, line in enumerate(recipient_lines):
        pdf.drawString(value_x, recipient_y - (idx * 8.2), line)

    envio_center_x = envio_x + ((left_x1 - envio_x) / 2)
    pdf.setFont("Helvetica-Bold", 7.8)
    pdf.drawCentredString(envio_center_x, box_top + 14, "Envio")
    pdf.setFont("Helvetica-Bold", 15.5)
    pdf.drawCentredString(envio_center_x, box_top - 2, row.shipment_number)

    meta_y = 27
    pdf.setFillColor(HexColor("#111827"))
    draw_wrapped(
        pdf,
        f"Contenido de la caja: {row.content or 'N/D'}",
        left_x0 + 8,
        meta_y,
        left_width - 16,
        font_name="Helvetica-Bold",
        font_size=7.3,
        leading=8.2,
        max_lines=1,
    )
    draw_wrapped(
        pdf,
        f"Piezas: {row.pieces or '1'}   Peso: {format_weight(row.weight_lb, row.weight_kg) or 'N/D'}",
        left_x0 + 8,
        meta_y - 11,
        left_width - 16,
        font_name="Helvetica-Bold",
        font_size=7.3,
        leading=8.2,
        max_lines=1,
    )
    draw_wrapped(
        pdf,
        f"Valor declarado: {row.declared_value or 'N/D'}   Posicion arancelaria: {row.tariff_code or 'N/D'}",
        left_x0 + 8,
        meta_y - 22,
        left_width - 16,
        font_name="Helvetica-Bold",
        font_size=7.3,
        leading=8.2,
        max_lines=1,
    )

    right_center = right_x0 + right_width / 2
    draw_centered_block(
        pdf,
        right_center,
        PAGE_HEIGHT - 26,
        [
            ("Al momento de recibir tu", "Helvetica", 8.5, "#6B7280"),
            ("envio esta etiqueta no debe", "Helvetica", 8.5, "#6B7280"),
            ("estar rota o rasgada", "Helvetica", 8.5, "#6B7280"),
        ],
        leading=10,
    )

    draw_warning_icon(pdf, right_center, PAGE_HEIGHT - 72, 18)
    draw_box_icon(pdf, right_x0 + 42, PAGE_HEIGHT - 122, 34)
    draw_house_icon(pdf, right_x1 - 42, PAGE_HEIGHT - 122, 34)
    pdf.setStrokeColor(HexColor("#9CA3AF"))
    pdf.setLineWidth(0.8)
    pdf.line(right_x0 + 82, PAGE_HEIGHT - 122, right_x1 - 82, PAGE_HEIGHT - 122)

    draw_centered_block(
        pdf,
        right_center,
        PAGE_HEIGHT - 154,
        [
            ("Compra Online", "Helvetica-Bold", 8.4, "#4B5563"),
            ("Personal Shopper", "Helvetica-Bold", 8.4, "#4B5563"),
            ("Recogida en tienda (Pick up)", "Helvetica-Bold", 8.4, "#4B5563"),
        ],
        leading=15,
    )

    attention_y = 54
    pdf.setFillColor(HexColor("#6B7280"))
    pdf.setFont("Helvetica", 8)
    pdf.drawCentredString(right_center, attention_y, "Linea unica de Atencion")
    pdf.setFont("Helvetica", 8.4)
    pdf.drawCentredString(right_center, attention_y - 15, "(318) 242-8086")
    pdf.setFont("Helvetica-Bold", 9)
    pdf.drawCentredString(right_center, attention_y - 41, "www.encargomio.com")


def generate_pdf(rows: list[LabelRow], output_path: Path) -> None:
    ensure_reportlab()
    pdf = canvas.Canvas(str(output_path), pagesize=PAGE_SIZE)

    draw_label(pdf, rows[0])
    pdf.save()


def generate_pdf_bytes(row: LabelRow) -> bytes:
    ensure_reportlab()
    buffer = io.BytesIO()
    pdf = canvas.Canvas(buffer, pagesize=PAGE_SIZE)
    draw_label(pdf, row)
    pdf.save()
    return buffer.getvalue()


def generate_pdfs(rows: list[LabelRow], output_dir: Path) -> list[Path]:
    ensure_reportlab()
    created_files: list[Path] = []
    used_names: dict[str, int] = {}

    for row in rows:
        base_name = sanitize_filename(row.shipment_number)
        duplicate_number = used_names.get(base_name, 0) + 1
        used_names[base_name] = duplicate_number

        if duplicate_number == 1:
            filename = f"{base_name}.pdf"
        else:
            filename = f"{base_name}_{duplicate_number}.pdf"

        output_path = output_dir / filename
        generate_pdf([row], output_path)
        created_files.append(output_path)

    return created_files


def generate_zip_bytes(rows: list[LabelRow]) -> bytes:
    ensure_reportlab()
    buffer = io.BytesIO()
    used_names: dict[str, int] = {}

    with zipfile.ZipFile(buffer, "w", compression=zipfile.ZIP_DEFLATED) as archive:
        for row in rows:
            base_name = sanitize_filename(row.shipment_number)
            duplicate_number = used_names.get(base_name, 0) + 1
            used_names[base_name] = duplicate_number

            if duplicate_number == 1:
                filename = f"{base_name}.pdf"
            else:
                filename = f"{base_name}_{duplicate_number}.pdf"

            archive.writestr(filename, generate_pdf_bytes(row))

    return buffer.getvalue()


# ---------------------------------------------------------------------------
# MOTOR DE TEMAS (etiquetas GENERICAS simuladas, SIN marca)
# ---------------------------------------------------------------------------
# Este bloque es independiente de la etiqueta con marca Encargomio (draw_label,
# generate_pdf_bytes, etc.). Aqui SOLO se define el catalogo de estilos; el
# dibujo de los temas se agrega en un paso posterior.
#
# Restricciones de diseno:
#   - Pagina 4x6 exactas: usar PAGE_WIDTH / PAGE_HEIGHT existentes.
#   - Solo fuentes built-in de reportlab (sin fuentes externas).
#   - Ningun texto de marca / logo / nombre de empresa: rotulos genericos.

# Fuentes estandar (built-in) de reportlab; no requieren registro previo.
THEME_FONT_FAMILIES = (
    "Helvetica",
    "Helvetica-Bold",
    "Courier",
    "Courier-Bold",
    "Times-Roman",
    "Times-Bold",
)

# Posiciones posibles para el bloque de codigo de barras.
THEME_BARCODE_POSITIONS = ("top", "top-left", "top-right", "mid")

# Colores de acento (lineas / titulos) en hex.
THEME_ACCENT_COLORS = (
    "#111827", "#F97316", "#2563EB", "#059669", "#DC2626", "#7C3AED",
    "#0891B2", "#CA8A04", "#DB2777", "#4B5563", "#0F766E", "#B91C1C",
    "#1D4ED8", "#15803D", "#9333EA",
)

# Colores para el marco exterior en hex.
THEME_BORDER_COLORS = ("#111827", "#6B7280", "#9CA3AF", "#374151", "#000000")

# Estilos de linea divisoria entre bloques.
THEME_DIVIDER_STYLES = ("solid", "dashed", "dotted", "double")

# Alineacion de los bloques de texto.
THEME_ALIGNMENTS = ("left", "center")

# Pares de rotulos GENERICOS (remitente / destinatario). Nunca marca.
THEME_LABEL_PAIRS = (
    ("De:", "Para:"),
    ("Remitente", "Destinatario"),
    ("From", "To"),
    ("Origen", "Destino"),
    ("Envia", "Recibe"),
    ("Sender", "Recipient"),
    ("Remite", "Recibe"),
)

# Bloques que se apilan verticalmente (su orden es parte de la variacion).
THEME_BASE_BLOCKS = ("barcode", "sender", "recipient", "content")


@dataclass(frozen=True)
class LabelTheme:
    """Parametros que definen un diseno de etiqueta generica 4x6.

    Un tema NO contiene datos del envio ni marca; solo describe el estilo y el
    layout. La funcion de dibujo (paso posterior) consume estos parametros.
    """

    theme_id: int
    # Codigo de barras
    barcode_position: str          # "top" | "top-left" | "top-right" | "mid"
    barcode_height: float          # alto del Code128 en puntos
    barcode_bar_width: float       # grosor de barra del Code128
    show_barcode_text: bool        # mostrar el valor legible bajo el codigo
    # Tipografia
    font_family: str               # familia base (built-in reportlab)
    title_size: float
    body_size: float
    label_size: float
    line_leading: float            # factor de interlineado (x body_size)
    uppercase_labels: bool         # rotulos en MAYUSCULAS
    # Marco / divisores
    has_outer_border: bool
    border_width: float
    border_color: str              # hex
    has_divider: bool
    divider_style: str             # "solid" | "dashed" | "dotted" | "double" | "none"
    # Color / composicion
    accent_color: str              # hex
    alignment: str                 # "left" | "center"
    padding: float                 # margen interno en puntos
    block_order: tuple[str, ...]   # orden vertical de los bloques
    # Rotulos genericos (sin marca)
    sender_label: str
    recipient_label: str

    def signature(self) -> str:
        """Firma visual (excluye theme_id) para detectar temas repetidos."""
        parts = [
            f"{name}={getattr(self, name)}"
            for name in self.__dataclass_fields__
            if name != "theme_id"
        ]
        return "|".join(parts)

    def describe(self) -> str:
        """Representacion legible multilinea para inspeccion."""
        lines = [
            f"Tema #{self.theme_id}",
            f"  barcode      : pos={self.barcode_position} "
            f"h={self.barcode_height} bar_w={self.barcode_bar_width} "
            f"texto={'si' if self.show_barcode_text else 'no'}",
            f"  tipografia   : {self.font_family} "
            f"(title={self.title_size} body={self.body_size} "
            f"label={self.label_size} leading={self.line_leading} "
            f"mayus={'si' if self.uppercase_labels else 'no'})",
            f"  marco        : "
            + (
                f"si w={self.border_width} color={self.border_color}"
                if self.has_outer_border
                else "no"
            ),
            f"  divisor      : "
            + (self.divider_style if self.has_divider else "no"),
            f"  acento       : {self.accent_color}",
            f"  alineacion   : {self.alignment}   padding={self.padding}",
            f"  orden bloques: {' > '.join(self.block_order)}",
            f"  rotulos      : '{self.sender_label}' / '{self.recipient_label}'",
        ]
        return "\n".join(lines)


def build_theme_catalog(n: int = 50) -> list[LabelTheme]:
    """Genera una lista DETERMINISTA de ``n`` temas visualmente distintos.

    La semilla interna (random.Random(42)) garantiza que el catalogo #0..#(n-1)
    sea siempre el mismo entre ejecuciones. Los temas repetidos (misma firma
    visual) se descartan para maximizar la variedad.
    """
    rng = random.Random(42)
    catalog: list[LabelTheme] = []
    seen: set[str] = set()
    max_attempts = n * 60

    for _ in range(max_attempts):
        if len(catalog) >= n:
            break

        block_order = list(THEME_BASE_BLOCKS)
        rng.shuffle(block_order)
        sender_label, recipient_label = rng.choice(THEME_LABEL_PAIRS)
        has_outer_border = rng.random() < 0.70
        has_divider = rng.random() < 0.75

        theme = LabelTheme(
            theme_id=len(catalog),
            barcode_position=rng.choice(THEME_BARCODE_POSITIONS),
            barcode_height=round(rng.uniform(42.0, 58.0), 1),
            barcode_bar_width=round(rng.uniform(1.40, 2.10), 2),
            show_barcode_text=rng.random() < 0.80,
            font_family=rng.choice(THEME_FONT_FAMILIES),
            title_size=round(rng.uniform(14.0, 22.0), 1),
            body_size=round(rng.uniform(7.0, 10.0), 1),
            label_size=round(rng.uniform(8.0, 11.0), 1),
            line_leading=round(rng.uniform(1.15, 1.50), 2),
            uppercase_labels=rng.random() < 0.35,
            has_outer_border=has_outer_border,
            border_width=round(rng.uniform(0.6, 1.6), 1) if has_outer_border else 0.0,
            border_color=rng.choice(THEME_BORDER_COLORS) if has_outer_border else "#FFFFFF",
            has_divider=has_divider,
            divider_style=rng.choice(THEME_DIVIDER_STYLES) if has_divider else "none",
            accent_color=rng.choice(THEME_ACCENT_COLORS),
            alignment=rng.choice(THEME_ALIGNMENTS),
            padding=round(rng.uniform(8.0, 18.0), 1),
            block_order=tuple(block_order),
            sender_label=sender_label,
            recipient_label=recipient_label,
        )

        signature = theme.signature()
        if signature in seen:
            continue
        seen.add(signature)
        catalog.append(theme)

    if len(catalog) < n:
        raise RuntimeError(
            f"Solo se generaron {len(catalog)} temas distintos de {n} solicitados."
        )

    return catalog


def pick_theme(catalog: list[LabelTheme], key: object) -> LabelTheme:
    """Devuelve un tema del catalogo de forma estable: mismo ``key`` -> mismo tema.

    Usa un hash estable (md5) sobre ``str(key)`` para no depender de la
    aleatorizacion del hash de Python entre procesos.
    """
    if not catalog:
        raise ValueError("El catalogo de temas esta vacio.")
    digest = hashlib.md5(str(key).encode("utf-8")).hexdigest()
    index = int(digest, 16) % len(catalog)
    return catalog[index]


# ---------------------------------------------------------------------------
# Fin del MOTOR DE TEMAS
# ---------------------------------------------------------------------------


# ---------------------------------------------------------------------------
# DIBUJO DE ETIQUETA GENERICA (consume un LabelTheme). SIN marca.
# ---------------------------------------------------------------------------
# Independiente de draw_label (etiqueta con marca Encargomio): esta funcion NO
# dibuja logos, web, telefonos, canales de compra ni el saludo "!Hola!".
#
# Estrategia para GARANTIZAR que todo cabe en 4x6 sin solaparse, en cualquier
# tema:
#   1. Layout de FLUJO vertical: los bloques se apilan segun theme.block_order,
#      cada uno debajo del anterior (nada de coordenadas absolutas fijas).
#   2. Los textos se envuelven/truncan con wrap_text dentro del ancho util.
#   3. Se mide la altura total y, si excede la pagina, se aplica una escala
#      global a fuentes y barcode y se vuelve a medir (fuentes grandes + mucho
#      texto => se encoge hasta caber).
#   4. Recorte duro de seguridad al dibujar: nunca se pinta por debajo del
#      margen inferior; si algo no cabe, se corta y se reporta truncado.
#
# Nota: draw_code128_barcode fija barWidth=1.2 internamente (ignora el ancho de
# barra del tema). Para honrar theme.barcode_bar_width y ademas CLAMPEAR el
# ancho del codigo a su region (que no se salga), se usa un dibujante propio.

GENERIC_TEXT_COLOR = "#111827"
GENERIC_BORDER_INSET = 3.0

# Rotulo generico que precede al numero de envio repetido en grande. Varia por
# tema (indexado por theme_id) sin tocar build_theme_catalog, para no alterar el
# mapeo estable guia -> diseno ya existente.
GENERIC_TRACKING_LABELS = ("Envio", "Guia", "Tracking", "Envío")

# Rotulos FIJOS de remitente/destinatario (como la etiqueta con marca). El texto
# es fijo en los 50 temas; el estilo (mayusculas, fuente, color) sigue el tema.
GENERIC_SENDER_LABEL = "Remitente"
GENERIC_RECIPIENT_LABEL = "Destinatario"

# Metricas para reservar la altura REAL del numero de envio grande y garantizar
# separacion con el codigo de barras (sin solapamiento).
GENERIC_BIG_ASCENT = 0.80    # alto sobre la linea base (x tamano de fuente)
GENERIC_BIG_DESCENT = 0.22   # bajo la linea base (x tamano de fuente)
GENERIC_BIG_BARCODE_GAP = 4.0  # aire fijo entre el numero grande y el barcode


def _generic_wrap_capped(
    text: str, font_name: str, font_size: float, max_width: float, max_lines: int
) -> tuple[list[str], bool]:
    """Envuelve el texto y lo limita a ``max_lines`` (con '...' si sobra)."""
    lines = wrap_text(text, font_name, font_size, max_width)
    if len(lines) <= max_lines:
        return lines, False
    lines = lines[:max_lines]
    last = lines[-1]
    while last and stringWidth(last + "...", font_name, font_size) > max_width:
        last = last[:-1]
    lines[-1] = (last.rstrip() + "...") if last else "..."
    return lines, True


def _draw_generic_barcode(
    pdf: "canvas.Canvas",
    value: str,
    region_x: float,
    region_width: float,
    y: float,
    height: float,
    bar_width: float,
) -> tuple[float, float]:
    """Dibuja un Code128 centrado dentro de [region_x, region_x+region_width].

    Respeta ``bar_width`` del tema y reduce el grosor si el codigo no cabe en la
    region, de modo que nunca se salga horizontalmente. Devuelve (x, ancho).
    """
    ensure_reportlab()
    bw = max(0.4, bar_width)
    drawing = createBarcodeDrawing(
        "Code128", value=value, barHeight=height, barWidth=bw, humanReadable=False
    )
    # El ancho del Code128 NO escala linealmente con barWidth (reportlab cuantiza
    # el grosor de barra), asi que un solo reescalado puede quedar unos puntos por
    # encima de la region. Iteramos con un pequeno margen hasta que quepa.
    guard = 0
    while drawing.width > region_width and drawing.width > 0 and guard < 8:
        bw = bw * (region_width / drawing.width) * 0.98
        drawing = createBarcodeDrawing(
            "Code128", value=value, barHeight=height, barWidth=bw, humanReadable=False
        )
        guard += 1
    draw_x = region_x + max(0.0, (region_width - drawing.width) / 2)
    renderPDF.draw(drawing, pdf, draw_x, y)
    return draw_x, drawing.width


def _draw_generic_divider(
    pdf: "canvas.Canvas", x0: float, x1: float, y: float, theme: LabelTheme
) -> None:
    """Dibuja una linea divisoria segun theme.divider_style."""
    pdf.setStrokeColor(HexColor(theme.accent_color))
    pdf.setLineWidth(0.7)
    style = theme.divider_style
    if style == "dashed":
        pdf.setDash(4, 2)
    elif style == "dotted":
        pdf.setDash(1, 2)
    else:
        pdf.setDash()
    if style == "double":
        pdf.setDash()
        pdf.line(x0, y + 1.0, x1, y + 1.0)
        pdf.line(x0, y - 1.0, x1, y - 1.0)
    else:
        pdf.line(x0, y, x1, y)
    pdf.setDash()


def draw_generic_label(pdf: "canvas.Canvas", row: LabelRow, theme: LabelTheme) -> dict:
    """Dibuja una etiqueta GENERICA (sin marca) en 4x6 usando ``theme``.

    Devuelve un reporte: {truncated, wrap_truncated, hard_truncated, scale,
    bbox:(min_x,min_y,max_x,max_y), content:(left,bottom,right,top)}.
    """
    ensure_reportlab()

    pad = float(theme.padding)
    content_left = pad
    content_right = PAGE_WIDTH - pad
    content_top = PAGE_HEIGHT - pad
    content_bottom = pad
    content_width = content_right - content_left
    available_height = content_top - content_bottom
    indent = 2.0
    center_x = (content_left + content_right) / 2.0
    font = theme.font_family
    leading = theme.line_leading

    # --- Construccion medible de bloques (para una escala dada) -------------
    def build_blocks(scale: float):
        body_sz = max(5.0, theme.body_size * scale)
        label_sz = max(6.0, theme.label_size * scale)
        text_width = content_width - 2 * indent
        wrap_trunc = False

        def line_h(sz: float) -> float:
            return sz * leading

        def text_item(text: str, sz: float, color: str, is_label: bool = False) -> dict:
            return {
                "kind": "text",
                "text": text,
                "size": sz,
                "color": color,
                "height": line_h(sz),
                "is_label": is_label,
            }

        def address_block(label_text, name, address, city, state):
            nonlocal wrap_trunc
            items = [
                text_item(
                    label_text.upper() if theme.uppercase_labels else label_text,
                    label_sz,
                    theme.accent_color,
                    is_label=True,
                )
            ]
            name_lines, t = _generic_wrap_capped(name or "N/D", font, body_sz, text_width, 1)
            wrap_trunc = wrap_trunc or t
            for ln in name_lines:
                items.append(text_item(ln, body_sz, GENERIC_TEXT_COLOR))
            if address:
                addr_lines, t = _generic_wrap_capped(address, font, body_sz, text_width, 2)
                wrap_trunc = wrap_trunc or t
                for ln in addr_lines:
                    items.append(text_item(ln, body_sz, GENERIC_TEXT_COLOR))
            city_state = " ".join(part for part in [city, state] if part)
            if city_state:
                cs_lines, t = _generic_wrap_capped(city_state, font, body_sz, text_width, 1)
                wrap_trunc = wrap_trunc or t
                for ln in cs_lines:
                    items.append(text_item(ln, body_sz, GENERIC_TEXT_COLOR))
            return items

        blocks: dict[str, list[dict]] = {}
        # Rotulos FIJOS Remitente/Destinatario (se ignora la variacion por tema de
        # estos dos rotulos; el estilo/mayusculas del tema si se respeta).
        blocks["sender"] = address_block(
            GENERIC_SENDER_LABEL, row.sender_name, row.sender_address,
            row.sender_city, row.sender_state,
        )
        blocks["recipient"] = address_block(
            GENERIC_RECIPIENT_LABEL, row.recipient_name, row.recipient_address,
            row.recipient_city, row.recipient_state,
        )

        content_items: list[dict] = []
        content_text = f"Contenido de la caja: {row.content or 'N/D'}"
        c_lines, t = _generic_wrap_capped(content_text, font, body_sz, text_width, 2)
        wrap_trunc = wrap_trunc or t
        for ln in c_lines:
            content_items.append(text_item(ln, body_sz, GENERIC_TEXT_COLOR))
        weight = format_weight(row.weight_lb, row.weight_kg) or "N/D"
        pw_text = f"Piezas: {row.pieces or '1'}   Peso: {weight}"
        pw_lines, t = _generic_wrap_capped(pw_text, font, body_sz, text_width, 1)
        wrap_trunc = wrap_trunc or t
        for ln in pw_lines:
            content_items.append(text_item(ln, body_sz, GENERIC_TEXT_COLOR))
        # Valor declarado + posicion arancelaria (mismo rotulo fijo que la etiqueta
        # con marca). Va DENTRO del bloque de contenido para no romper el flujo.
        vt_text = (
            f"Valor declarado: {row.declared_value or 'N/D'}   "
            f"Posicion arancelaria: {row.tariff_code or 'N/D'}"
        )
        vt_lines, t = _generic_wrap_capped(vt_text, font, body_sz, text_width, 1)
        wrap_trunc = wrap_trunc or t
        for ln in vt_lines:
            content_items.append(text_item(ln, body_sz, GENERIC_TEXT_COLOR))
        blocks["content"] = content_items

        bar_h = max(30.0, theme.barcode_height * scale)
        num_text = (row.shipment_number or "").strip() or "0"
        # Numero pequeno bajo el codigo (legible), a tamano de rotulo.
        num_size = max(6.0, label_sz)
        while num_size > 6.0 and stringWidth(num_text, font, num_size) > text_width:
            num_size -= 0.5
        show_text = theme.show_barcode_text
        barcode_total = bar_h + (2.0 + line_h(num_size) if show_text else 0.0)
        barcode_item = {
            "kind": "barcode",
            "value": num_text,
            "bar_height": bar_h,
            "bar_width": theme.barcode_bar_width,
            "show_text": show_text,
            "num_size": num_size,
            "gap": 2.0,
            "height": barcode_total,
        }
        # Numero de envio repetido en GRANDE (destacado), con rotulo generico que
        # varia por tema. Va como encabezado del bloque del barcode, dentro del
        # flujo, para que se mida/escale como los demas items.
        track_label = GENERIC_TRACKING_LABELS[theme.theme_id % len(GENERIC_TRACKING_LABELS)]
        if theme.uppercase_labels:
            track_label = track_label.upper()
        big_text = f"{track_label} {num_text}".strip()
        big_size = max(9.0, theme.title_size * 1.5 * scale)
        while big_size > 4.0 and stringWidth(big_text, font, big_size) > text_width:
            big_size -= 0.5
        # Altura reservada = alto real de la fuente (asc+desc) + gap fijo hacia el
        # barcode. Asi el barcode se ubica desde el borde inferior REAL del numero
        # grande + gap, garantizando que no se solapen en ningun tema/escala.
        big_item = {
            "kind": "bignum",
            "text": big_text,
            "size": big_size,
            "color": theme.accent_color,
            "height": big_size * (GENERIC_BIG_ASCENT + GENERIC_BIG_DESCENT) + GENERIC_BIG_BARCODE_GAP,
        }
        blocks["barcode"] = [big_item, barcode_item]

        gap = max(3.0, body_sz * 0.8)
        ordered = [(name, blocks[name]) for name in theme.block_order]
        total = 0.0
        for i, (_, items) in enumerate(ordered):
            total += sum(it["height"] for it in items)
            if i < len(ordered) - 1:
                total += gap
        return ordered, total, wrap_trunc, gap

    # --- Escala-para-caber ---------------------------------------------------
    # Se reserva un colchon inferior (para el descendente de la ultima linea)
    # de modo que la escala absorba el sobrante antes de tener que recortar.
    bottom_cushion = 6.0
    available_fit = available_height - bottom_cushion
    scale = 1.0
    ordered, total, wrap_trunc, gap = build_blocks(scale)
    for _ in range(80):
        if total <= available_fit or scale <= 0.4:
            break
        scale *= 0.95
        ordered, total, wrap_trunc, gap = build_blocks(scale)

    # --- Dibujo --------------------------------------------------------------
    extents: list[tuple[float, float, float, float]] = []

    def record(x0: float, y0: float, x1: float, y1: float) -> None:
        extents.append((x0, y0, x1, y1))

    pdf.setFillColor(HexColor("#FFFFFF"))
    pdf.rect(0, 0, PAGE_WIDTH, PAGE_HEIGHT, stroke=0, fill=1)

    if theme.has_outer_border:
        inset = GENERIC_BORDER_INSET
        pdf.setStrokeColor(HexColor(theme.border_color))
        pdf.setLineWidth(theme.border_width)
        pdf.rect(inset, inset, PAGE_WIDTH - 2 * inset, PAGE_HEIGHT - 2 * inset, stroke=1, fill=0)

    hard_trunc = False
    y_cursor = content_top
    big_bbox = None
    barcode_bbox = None

    for bi, (_, items) in enumerate(ordered):
        for it in items:
            if it["kind"] == "barcode":
                bar_h = it["bar_height"]
                if "left" in theme.barcode_position:
                    region_x, region_w = content_left, content_width * 0.6
                elif "right" in theme.barcode_position:
                    region_w = content_width * 0.6
                    region_x = content_right - region_w
                else:
                    region_x, region_w = content_left, content_width
                bar_bottom = y_cursor - bar_h
                if bar_bottom < content_bottom:
                    hard_trunc = True
                    break
                draw_x, drawn_w = _draw_generic_barcode(
                    pdf, it["value"], region_x, region_w, bar_bottom, bar_h, it["bar_width"]
                )
                record(draw_x, bar_bottom, draw_x + drawn_w, bar_bottom + bar_h)
                barcode_bbox = (draw_x, bar_bottom, draw_x + drawn_w, bar_bottom + bar_h)
                y_cursor = bar_bottom - it["gap"]
                if it["show_text"]:
                    ns = it["num_size"]
                    y_cursor -= ns * leading
                    if y_cursor - ns * 0.25 < content_bottom:
                        hard_trunc = True
                        break
                    txt = it["value"]
                    w = stringWidth(txt, font, ns)
                    pdf.setFillColor(HexColor(GENERIC_TEXT_COLOR))
                    pdf.setFont(font, ns)
                    if "left" in theme.barcode_position:
                        tx = content_left + indent
                        pdf.drawString(tx, y_cursor, txt)
                        record(tx, y_cursor - ns * 0.2, tx + w, y_cursor + ns * 0.8)
                    elif "right" in theme.barcode_position:
                        tx = content_right - indent
                        pdf.drawRightString(tx, y_cursor, txt)
                        record(tx - w, y_cursor - ns * 0.2, tx, y_cursor + ns * 0.8)
                    else:
                        pdf.drawCentredString(center_x, y_cursor, txt)
                        record(center_x - w / 2, y_cursor - ns * 0.2, center_x + w / 2, y_cursor + ns * 0.8)
            elif it["kind"] == "bignum":
                # Numero de envio grande: se reserva su alto real (asc/desc) y se
                # deja un gap fijo antes del barcode -> nunca se solapan.
                sz = it["size"]
                top = y_cursor
                baseline = top - GENERIC_BIG_ASCENT * sz
                bottom = baseline - GENERIC_BIG_DESCENT * sz
                if bottom - GENERIC_BIG_BARCODE_GAP < content_bottom:
                    hard_trunc = True
                    break
                txt = it["text"]
                w = stringWidth(txt, font, sz)
                pdf.setFillColor(HexColor(it["color"]))
                pdf.setFont(font, sz)
                if theme.alignment == "center":
                    pdf.drawCentredString(center_x, baseline, txt)
                    x0, x1 = center_x - w / 2, center_x + w / 2
                else:
                    x0 = content_left + indent
                    pdf.drawString(x0, baseline, txt)
                    x1 = x0 + w
                record(x0, bottom, x1, top)
                big_bbox = (x0, bottom, x1, top)
                # el barcode (siguiente item) arranca desde aca, con gap garantizado
                y_cursor = bottom - GENERIC_BIG_BARCODE_GAP
            else:
                sz = it["size"]
                y_cursor -= it["height"]
                if y_cursor - sz * 0.25 < content_bottom:
                    hard_trunc = True
                    break
                txt = it["text"]
                w = stringWidth(txt, font, sz)
                pdf.setFillColor(HexColor(it["color"]))
                pdf.setFont(font, sz)
                if theme.alignment == "center":
                    pdf.drawCentredString(center_x, y_cursor, txt)
                    x0, x1 = center_x - w / 2, center_x + w / 2
                else:
                    x0 = content_left + indent
                    pdf.drawString(x0, y_cursor, txt)
                    x1 = x0 + w
                record(x0, y_cursor - sz * 0.2, x1, y_cursor + sz * 0.8)
        if hard_trunc:
            break
        if bi < len(ordered) - 1:
            divider_y = y_cursor - gap / 2
            if theme.has_divider and divider_y > content_bottom:
                _draw_generic_divider(pdf, content_left, content_right, divider_y, theme)
                record(content_left, divider_y - 1.0, content_right, divider_y + 1.0)
            y_cursor -= gap

    if extents:
        min_x = min(e[0] for e in extents)
        min_y = min(e[1] for e in extents)
        max_x = max(e[2] for e in extents)
        max_y = max(e[3] for e in extents)
    else:
        min_x = min_y = max_x = max_y = 0.0

    # Datos del numero de envio repetido en grande (primer item de texto del
    # bloque del barcode), para inspeccion/pruebas.
    big_number_text = ""
    big_number_size = 0.0
    for name, items in ordered:
        if name == "barcode":
            for it in items:
                if it["kind"] == "bignum":
                    big_number_text = it["text"]
                    big_number_size = round(it["size"], 2)
                    break
            break

    return {
        "theme_id": theme.theme_id,
        "scale": round(scale, 3),
        "wrap_truncated": wrap_trunc,
        "hard_truncated": hard_trunc,
        "truncated": bool(wrap_trunc or hard_trunc),
        "big_number": big_number_text,
        "big_number_size": big_number_size,
        "body_size": round(max(5.0, theme.body_size * scale), 2),
        "big_bbox": big_bbox,
        "barcode_bbox": barcode_bbox,
        "bbox": (round(min_x, 2), round(min_y, 2), round(max_x, 2), round(max_y, 2)),
        "content": (content_left, content_bottom, content_right, content_top),
    }


def generate_generic_pdf_bytes(row: LabelRow, theme: LabelTheme) -> bytes:
    """Genera los bytes de un PDF 4x6 con una etiqueta GENERICA segun ``theme``."""
    ensure_reportlab()
    buffer = io.BytesIO()
    pdf = canvas.Canvas(buffer, pagesize=PAGE_SIZE)
    draw_generic_label(pdf, row, theme)
    pdf.save()
    return buffer.getvalue()


def generate_generic_zip_bytes(
    rows: list[LabelRow],
    catalog: list[LabelTheme] | None = None,
) -> tuple[bytes, list[dict]]:
    """Arma un ZIP con una etiqueta GENERICA (sin marca) por fila.

    Paralela a generate_zip_bytes, pero:
      - Si ``catalog`` es None, lo construye con build_theme_catalog(50).
      - A cada fila le asigna un tema ESTABLE por guia con pick_theme, asi la
        misma guia siempre cae en el mismo diseno (reproducible).
      - Misma logica de nombres/duplicados que generate_zip_bytes (sufijos
        _2, _3, ... para guias repetidas).
      - Tolerante a errores por fila: si una fila falla al renderizar, se
        registra y se continua; NO tumba todo el ZIP.

    Devuelve ``(zip_bytes, failures)`` donde ``failures`` es una lista de
    ``{"shipment_number", "filename", "error"}`` por cada fila que fallo.
    """
    ensure_reportlab()
    if catalog is None:
        catalog = build_theme_catalog(50)

    buffer = io.BytesIO()
    used_names: dict[str, int] = {}
    failures: list[dict] = []

    with zipfile.ZipFile(buffer, "w", compression=zipfile.ZIP_DEFLATED) as archive:
        for row in rows:
            base_name = sanitize_filename(row.shipment_number)
            duplicate_number = used_names.get(base_name, 0) + 1
            used_names[base_name] = duplicate_number

            if duplicate_number == 1:
                filename = f"{base_name}.pdf"
            else:
                filename = f"{base_name}_{duplicate_number}.pdf"

            try:
                theme = pick_theme(catalog, row.shipment_number)
                archive.writestr(filename, generate_generic_pdf_bytes(row, theme))
            except Exception as exc:  # tolerante por fila: una fila mala no rompe el ZIP
                failures.append({
                    "shipment_number": row.shipment_number,
                    "filename": filename,
                    "error": f"{type(exc).__name__}: {exc}",
                })
                continue

    return buffer.getvalue(), failures


# ---------------------------------------------------------------------------
# Fin del DIBUJO GENERICO
# ---------------------------------------------------------------------------


def build_parser() -> argparse.ArgumentParser:
    parser = argparse.ArgumentParser(
        description="Genera un PDF de etiquetas 4x6 pulgadas a partir de un archivo Excel."
    )
    parser.add_argument("input", type=Path, help="Ruta del archivo Excel de entrada.")
    parser.add_argument(
        "-o",
        "--output",
        type=Path,
        default=Path("labels_4x6"),
        help="Carpeta de salida donde se guardara un PDF por fila. Default: labels_4x6",
    )
    parser.add_argument(
        "-s",
        "--sheet",
        type=str,
        default=None,
        help="Nombre de la hoja a procesar. Si no se indica, usa la primera hoja.",
    )
    return parser


def main() -> None:
    parser = build_parser()
    args = parser.parse_args()

    input_path = args.input.expanduser().resolve()
    output_dir = args.output.expanduser().resolve()

    if not input_path.exists():
        raise FileNotFoundError(f"No existe el archivo: {input_path}")

    rows = read_rows(input_path, args.sheet)
    if not rows:
        raise ValueError("No se encontraron filas con numero de envio/guia.")

    output_dir.mkdir(parents=True, exist_ok=True)
    created_files = generate_pdfs(rows, output_dir)
    print(f"Carpeta de salida: {output_dir}")
    print(f"Etiquetas creadas: {len(created_files)}")


if __name__ == "__main__":
    main()
